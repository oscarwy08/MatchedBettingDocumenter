from decimal import Decimal
from pathlib import Path

from sqlalchemy import select

from app.calculator import calculate
from app.db import init_db
from app.excel import sync_workbook
from app.models import Account, AccountType, Bet, BetStatus, BetType, Offer, Transfer, TransferKind
from app.seed import seed_accounts
from app.services import account_snapshot, offer_snapshot


def test_welcome_offer_balances_and_workbook(tmp_path: Path):
    db_path = tmp_path / "test.db"
    xlsx_path = tmp_path / "matched_betting.xlsx"
    Session = init_db(db_path)
    session = Session()
    seed_accounts(session)

    betfred = session.scalars(select(Account).where(Account.name == "Betfred")).one()
    smarkets = session.scalars(select(Account).where(Account.name == "Smarkets")).one()
    assert betfred.type == AccountType.BOOKIE
    assert smarkets.commission_percent == Decimal("2")

    session.add(Transfer(account_id=betfred.id, kind=TransferKind.DEPOSIT, amount=Decimal("10.00")))
    session.add(Transfer(account_id=smarkets.id, kind=TransferKind.DEPOSIT, amount=Decimal("50.00")))

    offer = Offer(
        name="Betfred Bet £10 Get £50",
        type="welcome",
        bookie_id=betfred.id,
        deposit_amount=Decimal("10.00"),
        free_funds=Decimal("50.00"),
    )
    session.add(offer)
    session.flush()

    qb = calculate(
        BetType.QUALIFYING,
        back_stake=10,
        back_odds=2,
        lay_odds="2.10",
        commission_percent=2,
    )
    qualifier = Bet(
        offer_id=offer.id,
        event="Liverpool vs Chelsea",
        market="Match odds / Liverpool",
        bet_type=BetType.QUALIFYING,
        bookie_id=betfred.id,
        exchange_id=smarkets.id,
        back_stake=qb.back_stake,
        back_odds=qb.back_odds,
        lay_stake=qb.lay_stake,
        lay_odds=qb.lay_odds,
        commission_percent=qb.commission_percent,
        cashback=Decimal("0"),
        liability=qb.liability,
        expected_profit=qb.expected_profit,
        expected_bookie_back=qb.if_back_wins.bookie,
        expected_exchange_back=qb.if_back_wins.exchange,
        expected_bookie_lay=qb.if_lay_wins.bookie,
        expected_exchange_lay=qb.if_lay_wins.exchange,
        status=BetStatus.LAY_WON,
        actual_bookie_profit=qb.if_lay_wins.bookie,
        actual_exchange_profit=qb.if_lay_wins.exchange,
        actual_profit=qb.if_lay_wins.total,
    )
    session.add(qualifier)

    fb = calculate(
        BetType.FREE_BET_SNR,
        back_stake=50,
        back_odds="5.50",
        lay_odds="5.20",
        commission_percent=2,
    )
    conversion = Bet(
        offer_id=offer.id,
        event="Arsenal vs Spurs",
        bet_type=BetType.FREE_BET_SNR,
        bookie_id=betfred.id,
        exchange_id=smarkets.id,
        back_stake=fb.back_stake,
        back_odds=fb.back_odds,
        lay_stake=fb.lay_stake,
        lay_odds=fb.lay_odds,
        commission_percent=fb.commission_percent,
        cashback=Decimal("0"),
        liability=fb.liability,
        expected_profit=fb.expected_profit,
        expected_bookie_back=fb.if_back_wins.bookie,
        expected_exchange_back=fb.if_back_wins.exchange,
        expected_bookie_lay=fb.if_lay_wins.bookie,
        expected_exchange_lay=fb.if_lay_wins.exchange,
        status=BetStatus.BACK_WON,
        actual_bookie_profit=fb.if_back_wins.bookie,
        actual_exchange_profit=fb.if_back_wins.exchange,
        actual_profit=fb.if_back_wins.total,
    )
    session.add(conversion)
    session.commit()
    session.refresh(offer)

    snap = offer_snapshot(offer)
    assert snap["status"] == "Used"
    assert snap["free_funds_used"] == Decimal("50.00")
    assert snap["free_funds"] == Decimal("50.00")
    assert snap["net_profit"] == qualifier.actual_profit + conversion.actual_profit
    assert snap["net_profit"] > 0

    bookie = account_snapshot(session, betfred)
    exchange = account_snapshot(session, smarkets)
    assert bookie["deposited"] == Decimal("10.00")
    assert bookie["balance"] == Decimal("10.00") + bookie["bookie_profit"]
    assert exchange["deposited"] == Decimal("50.00")
    assert exchange["balance"] == Decimal("50.00") + exchange["exchange_profit"]
    assert bookie["net_profit"] == snap["net_profit"]

    path = sync_workbook(session, xlsx_path)
    assert path.exists()
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert wb.sheetnames == ["Dashboard", "Offers", "Bets", "Accounts", "Transfers", "Tasks"]
    assert wb["Offers"]["A2"].value == "Betfred Bet £10 Get £50"
    assert wb["Bets"]["A1"].value == "Date"
    assert wb["Bets"]["B1"].value == "Placed"
    assert wb["Bets"]["C1"].value == "Settled"
    assert wb["Bets"]["E2"].value in {"Arsenal vs Spurs", "Liverpool vs Chelsea"}

    from app.excel import preview_workbook

    preview = preview_workbook(path)
    assert [sheet["name"] for sheet in preview] == [
        "Dashboard",
        "Offers",
        "Bets",
        "Accounts",
        "Transfers",
        "Tasks",
    ]
    offer_sheet = next(sheet for sheet in preview if sheet["name"] == "Offers")
    assert any("Betfred Bet £10 Get £50" in cell["value"] for row in offer_sheet["rows"] for cell in row)
    session.close()


def test_free_funds_status_from_conversions(tmp_path: Path):
    Session = init_db(tmp_path / "test.db")
    session = Session()
    seed_accounts(session)
    sky = session.scalars(select(Account).where(Account.name == "Sky Bet")).one()
    smarkets = session.scalars(select(Account).where(Account.name == "Smarkets")).one()
    offer = Offer(
        name="Bet £10 Get £30",
        type="welcome",
        bookie_id=sky.id,
        free_funds=Decimal("30.00"),
    )
    session.add(offer)
    session.flush()
    session.add(
        Bet(
            offer_id=offer.id,
            event="Qualifier",
            bet_type=BetType.QUALIFYING,
            bookie_id=sky.id,
            exchange_id=smarkets.id,
            back_stake=Decimal("10"),
            back_odds=Decimal("2"),
            lay_stake=Decimal("9.62"),
            lay_odds=Decimal("2.1"),
            commission_percent=Decimal("2"),
            cashback=Decimal("0"),
            liability=Decimal("10.58"),
            expected_profit=Decimal("-0.58"),
            expected_bookie_back=Decimal("10"),
            expected_exchange_back=Decimal("-10.58"),
            expected_bookie_lay=Decimal("-10"),
            expected_exchange_lay=Decimal("9.43"),
            status=BetStatus.LAY_WON,
        )
    )
    session.flush()
    session.refresh(offer)
    snap = offer_snapshot(offer)
    assert snap["status"] == "In progress"
    assert snap["free_funds_used"] == Decimal("0.00")

    session.add(
        Bet(
            offer_id=offer.id,
            event="Conversion",
            bet_type=BetType.FREE_BET_SNR,
            bookie_id=sky.id,
            exchange_id=smarkets.id,
            back_stake=Decimal("30"),
            back_odds=Decimal("5.5"),
            lay_stake=Decimal("25"),
            lay_odds=Decimal("5.2"),
            commission_percent=Decimal("2"),
            cashback=Decimal("0"),
            liability=Decimal("105"),
            expected_profit=Decimal("5"),
            expected_bookie_back=Decimal("0"),
            expected_exchange_back=Decimal("0"),
            expected_bookie_lay=Decimal("0"),
            expected_exchange_lay=Decimal("0"),
            status=BetStatus.PENDING,
        )
    )
    session.flush()
    session.refresh(offer)
    snap = offer_snapshot(offer)
    assert snap["free_funds_used"] == Decimal("30.00")
    assert snap["status"] == "Used"
    session.close()


def test_voided_free_bet_drops_used_funds_when_returned(tmp_path: Path):
    Session = init_db(tmp_path / "test.db")
    session = Session()
    seed_accounts(session)
    sky = session.scalars(select(Account).where(Account.name == "Sky Bet")).one()
    smarkets = session.scalars(select(Account).where(Account.name == "Smarkets")).one()
    offer = Offer(
        name="Get £30",
        type="welcome",
        bookie_id=sky.id,
        free_funds=Decimal("30.00"),
    )
    session.add(offer)
    session.flush()
    kept = Bet(
        offer_id=offer.id,
        event="Kept",
        bet_type=BetType.FREE_BET_SNR,
        bookie_id=sky.id,
        exchange_id=smarkets.id,
        back_stake=Decimal("20"),
        back_odds=Decimal("5"),
        lay_stake=Decimal("16"),
        lay_odds=Decimal("5.2"),
        commission_percent=Decimal("2"),
        cashback=Decimal("0"),
        liability=Decimal("67"),
        expected_profit=Decimal("4"),
        expected_bookie_back=Decimal("0"),
        expected_exchange_back=Decimal("0"),
        expected_bookie_lay=Decimal("0"),
        expected_exchange_lay=Decimal("0"),
        status=BetStatus.PENDING,
    )
    voided = Bet(
        offer_id=offer.id,
        event="Voided ten",
        bet_type=BetType.FREE_BET_SNR,
        bookie_id=sky.id,
        exchange_id=smarkets.id,
        back_stake=Decimal("10"),
        back_odds=Decimal("4"),
        lay_stake=Decimal("8"),
        lay_odds=Decimal("4.2"),
        commission_percent=Decimal("2"),
        cashback=Decimal("0"),
        liability=Decimal("25"),
        expected_profit=Decimal("2"),
        expected_bookie_back=Decimal("0"),
        expected_exchange_back=Decimal("0"),
        expected_bookie_lay=Decimal("0"),
        expected_exchange_lay=Decimal("0"),
        status=BetStatus.VOID,
        free_bet_returned=True,
    )
    session.add_all([kept, voided])
    session.flush()
    session.refresh(offer)
    snap = offer_snapshot(offer)
    assert snap["free_funds_used"] == Decimal("20.00")
    assert snap["status"] == "In progress"
    voided.free_bet_returned = False
    session.flush()
    session.refresh(offer)
    snap = offer_snapshot(offer)
    assert snap["free_funds_used"] == Decimal("30.00")
    assert snap["status"] == "Used"
    session.close()
