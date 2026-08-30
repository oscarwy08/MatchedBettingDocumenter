from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app import EXCEL_PATH
from app.dates import format_uk_time
from app.models import Account, AccountTask, AccountType, Bet, Offer, ScheduleEvent, Transfer
from app.services import account_snapshot, dashboard_stats, offer_snapshot

HEADER_FILL = PatternFill("solid", fgColor="1A2332")
HEADER_FONT = Font(name="Calibri", bold=True, color="F4F1EA")
MONEY_FORMAT = '£#,##0.00;£-#,##0.00;"–"'
THIN = Border(
    left=Side(style="thin", color="D9D3C7"),
    right=Side(style="thin", color="D9D3C7"),
    top=Side(style="thin", color="D9D3C7"),
    bottom=Side(style="thin", color="D9D3C7"),
)
ZEBRA = PatternFill("solid", fgColor="F7F4EE")
PROFIT_FONT = Font(name="Calibri", color="1F7A4D")
LOSS_FONT = Font(name="Calibri", color="B42318")
LABEL_FONT = Font(name="Calibri", bold=True, color="3F3A32")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1A2332")
SUB_FONT = Font(name="Calibri", size=11, color="6B6458")


def _money(cell, value) -> None:
    cell.value = float(value) if value is not None else 0
    cell.number_format = MONEY_FORMAT
    cell.alignment = Alignment(horizontal="right")
    try:
        amount = Decimal(str(value or 0))
        if amount > 0:
            cell.font = PROFIT_FONT
        elif amount < 0:
            cell.font = LOSS_FONT
    except Exception:
        pass


def _headers(ws, titles: list[str], row: int = 1) -> None:
    for col, title in enumerate(titles, start=1):
        cell = ws.cell(row, col, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN


def _autosize(ws, min_width: int = 12, max_width: int = 36) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = min_width
        for cell in column:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _stripe(ws, start_row: int, cols: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_col=cols):
        if row[0].row % 2 == 0:
            for cell in row:
                if cell.fill.fgColor is None or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = ZEBRA
        for cell in row:
            cell.border = THIN


def sync_workbook(session: Session, path: Path | None = None) -> Path:
    target = Path(path) if path else EXCEL_PATH
    target.parent.mkdir(parents=True, exist_ok=True)

    stats = dashboard_stats(session)
    offers = list(session.scalars(select(Offer).order_by(Offer.created_at.desc())))
    bets = list(session.scalars(select(Bet).order_by(Bet.date_placed.desc(), Bet.id.desc())))
    accounts = list(
        session.scalars(select(Account).order_by(Account.type, Account.name))
    )
    transfers = list(
        session.scalars(select(Transfer).order_by(Transfer.date.desc(), Transfer.id.desc()))
    )
    tasks = list(
        session.scalars(
            select(AccountTask)
            .options(selectinload(AccountTask.account))
            .order_by(AccountTask.due_on, AccountTask.id)
        )
    )

    wb = Workbook()

    dash = wb.active
    dash.title = "Dashboard"
    _write_dashboard(dash, stats)

    offers_ws = wb.create_sheet("Offers")
    _write_offers(offers_ws, offers)

    bets_ws = wb.create_sheet("Bets")
    _write_bets(bets_ws, bets)

    accounts_ws = wb.create_sheet("Accounts")
    _write_accounts(accounts_ws, [account_snapshot(session, a) for a in accounts])

    transfers_ws = wb.create_sheet("Transfers")
    _write_transfers(transfers_ws, transfers)

    events = list(
        session.scalars(
            select(ScheduleEvent)
            .options(selectinload(ScheduleEvent.bookie))
            .order_by(ScheduleEvent.due_on, ScheduleEvent.id)
        )
    )

    tasks_ws = wb.create_sheet("Tasks")
    _write_tasks(tasks_ws, tasks)

    calendar_ws = wb.create_sheet("Calendar")
    _write_calendar(calendar_ws, events)

    wb.save(target)
    return target


def _write_dashboard(ws, stats: dict) -> None:
    ws["A1"] = "Matched Betting Documenter"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws["A2"] = f"Updated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = SUB_FONT

    labels = [
        ("Total net profit (settled)", stats["net_profit"]),
        ("Pending expected profit", stats["pending_expected"]),
        ("Bankroll (all accounts)", stats["bankroll"]),
        ("Open lay liability", stats["open_liability"]),
        ("Profit this month", stats["month_profit"]),
        ("Pending bets", stats["pending_count"]),
        ("Settled bets", stats["settled_count"]),
    ]
    _headers(ws, ["Metric", "Value"], row=4)
    for i, (label, value) in enumerate(labels, start=5):
        ws.cell(i, 1, label).font = LABEL_FONT
        if isinstance(value, int):
            ws.cell(i, 2, value)
        else:
            _money(ws.cell(i, 2), value)

    ws["A13"] = "Net profit by bookie"
    ws["A13"].font = LABEL_FONT
    _headers(ws, ["Bookie", "Deposited", "Bookie profit", "Exchange profit", "Net profit", "Balance"], row=14)
    row = 15
    for snap in stats["profit_by_bookie"]:
        ws.cell(row, 1, snap["account"].name)
        _money(ws.cell(row, 2), snap["deposited"])
        _money(ws.cell(row, 3), snap["bookie_profit"])
        _money(ws.cell(row, 4), snap["exchange_profit"])
        _money(ws.cell(row, 5), snap["net_profit"])
        _money(ws.cell(row, 6), snap["balance"])
        row += 1
    if row == 15:
        ws.cell(row, 1, "No settled bookie profit yet")
    _autosize(ws)


def _write_offers(ws, offers: list[Offer]) -> None:
    _headers(
        ws,
        [
            "Offer",
            "Bookie",
            "Type",
            "Repeats",
            "Reload stake",
            "Reward",
            "Next due",
            "Deposited",
            "Free funds",
            "Bookie profit",
            "Exchange profit",
            "Net profit",
            "Status",
            "Legs",
            "Notes",
        ],
    )
    for i, offer in enumerate(offers, start=2):
        snap = offer_snapshot(offer)
        ws.cell(i, 1, offer.name)
        ws.cell(i, 2, offer.bookie.name)
        ws.cell(i, 3, offer.type.replace("_", " ").title())
        ws.cell(i, 4, (offer.reload_frequency or "").replace("_", " ").title())
        _money(ws.cell(i, 5), snap["reload_stake"])
        _money(ws.cell(i, 6), snap["reload_reward"])
        ws.cell(i, 7, offer.next_reload_on.isoformat() if offer.next_reload_on else "")
        _money(ws.cell(i, 8), snap["deposited"])
        _money(ws.cell(i, 9), snap["free_funds"])
        _money(ws.cell(i, 10), snap["bookie_profit"])
        _money(ws.cell(i, 11), snap["exchange_profit"])
        _money(ws.cell(i, 12), snap["net_profit"])
        ws.cell(i, 13, snap["status"])
        ws.cell(i, 14, snap["leg_count"])
        ws.cell(i, 15, offer.notes)
    _stripe(ws, 2, 15)
    _autosize(ws)


def _write_bets(ws, bets: list[Bet]) -> None:
    _headers(
        ws,
        [
            "Date",
            "Placed",
            "Settled",
            "Offer",
            "Event",
            "Market",
            "Type",
            "Bookie",
            "Back stake",
            "Back odds",
            "Exchange",
            "Lay stake",
            "Lay odds",
            "Commission %",
            "Liability",
            "Expected profit",
            "Actual profit",
            "Bookie P&L",
            "Exchange P&L",
            "Status",
            "Notes",
        ],
    )
    for i, bet in enumerate(bets, start=2):
        ws.cell(i, 1, bet.date_placed.isoformat())
        placed = format_uk_time(bet.placed_at or bet.date_placed)
        ws.cell(i, 2, "" if placed == "–" else placed)
        settled = format_uk_time(bet.settled_at)
        ws.cell(i, 3, "" if settled == "–" else settled)
        ws.cell(i, 4, bet.offer.name if bet.offer else "")
        ws.cell(i, 5, bet.event)
        ws.cell(i, 6, bet.market)
        ws.cell(i, 7, bet.bet_type.replace("_", " ").title())
        ws.cell(i, 8, bet.bookie.name)
        _money(ws.cell(i, 9), bet.back_stake)
        ws.cell(i, 10, float(bet.back_odds))
        ws.cell(i, 11, bet.exchange.name)
        _money(ws.cell(i, 12), bet.lay_stake)
        ws.cell(i, 13, float(bet.lay_odds))
        ws.cell(i, 14, float(bet.commission_percent))
        _money(ws.cell(i, 15), bet.liability)
        _money(ws.cell(i, 16), bet.expected_profit)
        _money(ws.cell(i, 17), bet.actual_profit)
        _money(ws.cell(i, 18), bet.actual_bookie_profit)
        _money(ws.cell(i, 19), bet.actual_exchange_profit)
        ws.cell(i, 20, bet.status.replace("_", " ").title())
        ws.cell(i, 21, bet.notes)
    _stripe(ws, 2, 21)
    _autosize(ws)


def _write_accounts(ws, snapshots: list[dict]) -> None:
    _headers(
        ws,
        [
            "Account",
            "Type",
            "Deposited",
            "Withdrawn",
            "Bookie profit",
            "Exchange profit",
            "Net profit",
            "Balance",
            "Commission %",
            "Last checked",
            "Priority",
            "Restriction",
            "Notes",
        ],
    )
    row = 2
    for snap in snapshots:
        account = snap["account"]
        used = (
            snap["deposited"]
            or snap["withdrawn"]
            or snap["net_profit"]
            or snap["balance"]
            or account.type == AccountType.EXCHANGE
        )
        if not used:
            continue
        ws.cell(row, 1, account.name)
        ws.cell(row, 2, account.type.title())
        _money(ws.cell(row, 3), snap["deposited"])
        _money(ws.cell(row, 4), snap["withdrawn"])
        _money(ws.cell(row, 5), snap["bookie_profit"])
        _money(ws.cell(row, 6), snap["exchange_profit"])
        _money(ws.cell(row, 7), snap["net_profit"])
        _money(ws.cell(row, 8), snap["balance"])
        ws.cell(row, 9, float(account.commission_percent))
        ws.cell(row, 10, account.last_checked_on.isoformat() if account.last_checked_on else "")
        ws.cell(row, 11, "Yes" if account.priority else "")
        ws.cell(row, 12, (account.restriction or "").replace("_", " ").title())
        ws.cell(row, 13, account.notes or "")
        row += 1
    _stripe(ws, 2, 13)
    _autosize(ws)


def _write_transfers(ws, transfers: list[Transfer]) -> None:
    _headers(ws, ["Date", "Account", "Kind", "Amount", "Offer", "Notes"])
    for i, transfer in enumerate(transfers, start=2):
        ws.cell(i, 1, transfer.date.isoformat())
        ws.cell(i, 2, transfer.account.name)
        ws.cell(i, 3, transfer.kind.title())
        _money(ws.cell(i, 4), transfer.amount)
        ws.cell(i, 5, transfer.offer.name if transfer.offer else "")
        ws.cell(i, 6, transfer.notes)
    _stripe(ws, 2, 6)
    _autosize(ws)


def _write_tasks(ws, tasks: list[AccountTask]) -> None:
    _headers(ws, ["Due", "Account", "Note", "Done"])
    for i, task in enumerate(tasks, start=2):
        ws.cell(i, 1, task.due_on.isoformat())
        ws.cell(i, 2, task.account.name if task.account else "")
        ws.cell(i, 3, task.note)
        ws.cell(i, 4, "Yes" if task.done else "")
    _stripe(ws, 2, 4)
    _autosize(ws)


def _write_calendar(ws, events: list[ScheduleEvent]) -> None:
    _headers(ws, ["Due", "Title", "Bookie", "Repeats", "Notes", "Done"])
    for i, event in enumerate(events, start=2):
        ws.cell(i, 1, event.due_on.isoformat())
        ws.cell(i, 2, event.title)
        ws.cell(i, 3, event.bookie.name if event.bookie else "")
        ws.cell(i, 4, event.repeat or "Once")
        ws.cell(i, 5, event.notes)
        ws.cell(i, 6, "Yes" if event.done else "")
    _stripe(ws, 2, 6)
    _autosize(ws)


def _preview_value(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    fmt = str(cell.number_format or "")
    if isinstance(value, (int, float, Decimal)) and ("£" in fmt or "#,##0.00" in fmt):
        amount = Decimal(str(value))
        sign = "−" if amount < 0 else ""
        return f"{sign}£{abs(amount):,.2f}"
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return str(value)
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def _is_header_cell(cell) -> bool:
    try:
        rgb = str(getattr(cell.fill.fgColor, "rgb", "") or "")
        return "1A2332" in rgb.upper()
    except Exception:
        return False


def _cell_pnl(cell) -> str:
    value = cell.value
    if not isinstance(value, (int, float, Decimal)):
        return ""
    amount = Decimal(str(value))
    fmt = str(cell.number_format or "")
    if "£" not in fmt and "#,##0.00" not in fmt:
        return ""
    if amount > 0:
        return "pnl-pos"
    if amount < 0:
        return "pnl-neg"
    return ""


def preview_workbook(path: Path | None = None) -> list[dict]:
    target = Path(path) if path else EXCEL_PATH
    if not target.exists():
        return []
    workbook = load_workbook(target, data_only=True)
    sheets: list[dict] = []
    for name in workbook.sheetnames:
        ws = workbook[name]
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            cells = [
                {
                    "value": _preview_value(cell),
                    "header": _is_header_cell(cell),
                    "pnl": _cell_pnl(cell),
                    "money": "£" in str(cell.number_format or "")
                    or "#,##0.00" in str(cell.number_format or ""),
                }
                for cell in row
            ]
            if any(cell["value"] for cell in cells):
                rows.append(cells)
        sheets.append(
            {
                "name": name,
                "slug": name.lower().replace(" ", "-"),
                "col_letters": [get_column_letter(i) for i in range(1, max_col + 1)],
                "rows": rows,
            }
        )
    return sheets
