"""Import bets, offers, accounts and transfers from an existing .xlsx workbook."""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.dates import parse_uk
from app.services import reconcile_offer_deposits
from app.models import (
    Account,
    AccountType,
    Bet,
    BetStatus,
    BetType,
    Offer,
    OfferType,
    Transfer,
    TransferKind,
)

ZERO = Decimal("0.00")

BET_ALIASES = {
    "date": {"date", "date of bet"},
    "placed_at": {"placed", "placed at", "time placed"},
    "settled_at": {"settled", "settled at", "time settled"},
    "offer": {"offer", "campaign", "promotion"},
    "event": {"event", "selection", "fixture"},
    "market": {"market"},
    "bet_type": {"type", "bet type", "offer type"},
    "bookie": {"bookie", "sportsbook", "bookmaker", "back bookie"},
    "back_stake": {"back stake", "qualifying stake", "stake", "free bet"},
    "back_odds": {"back odds", "odds"},
    "exchange": {"exchange", "exchange used"},
    "lay_stake": {"lay stake"},
    "lay_odds": {"lay odds"},
    "commission": {"commission", "commission %", "lay commission"},
    "liability": {"liability", "lay liability"},
    "expected": {"expected profit", "expected"},
    "actual": {"actual profit", "net profit", "profit"},
    "bookie_pl": {"bookie p l", "bookie pnl", "bookie profit"},
    "exchange_pl": {"exchange p l", "exchange pnl", "exchange profit"},
    "status": {"status"},
    "notes": {"notes", "note"},
}

ACCOUNT_ALIASES = {
    "name": {"account", "name", "bookie", "exchange"},
    "type": {"type"},
    "deposited": {"deposited", "deposit", "deposits"},
    "withdrawn": {"withdrawn", "withdrawal", "withdrawals"},
    "commission": {"commission", "commission %"},
}

OFFER_ALIASES = {
    "name": {"offer", "name"},
    "bookie": {"bookie", "sportsbook"},
    "type": {"type"},
    "deposited": {"deposited", "deposit"},
    "free_funds": {"free funds", "free bet", "free bets"},
    "notes": {"notes"},
}

TRANSFER_ALIASES = {
    "date": {"date"},
    "account": {"account"},
    "kind": {"kind", "type"},
    "amount": {"amount"},
    "offer": {"offer"},
    "notes": {"notes"},
}

BET_TYPE_MAP = {
    "qualifying": BetType.QUALIFYING,
    "qualifying bet": BetType.QUALIFYING,
    "qualifier": BetType.QUALIFYING,
    "free bet": BetType.FREE_BET_SNR,
    "free bet snr": BetType.FREE_BET_SNR,
    "free bet stake not returned": BetType.FREE_BET_SNR,
    "snr": BetType.FREE_BET_SNR,
    "free bet sr": BetType.FREE_BET_SR,
    "sr": BetType.FREE_BET_SR,
    "money back": BetType.MONEY_BACK,
    "money back if bet loses": BetType.MONEY_BACK,
    "other": BetType.OTHER,
    "other manual": BetType.OTHER,
    "normal": BetType.NORMAL,
    "normal unmatched": BetType.NORMAL,
    "unmatched": BetType.NORMAL,
    "acca": BetType.ACCA,
    "accumulator": BetType.ACCA,
    "builder": BetType.BUILDER,
    "bet builder": BetType.BUILDER,
    "mug": BetType.MUG,
    "mug bet": BetType.MUG,
}

STATUS_MAP = {
    "pending": BetStatus.PENDING,
    "back won": BetStatus.BACK_WON,
    "lay won": BetStatus.LAY_WON,
    "void": BetStatus.VOID,
    "void pushed": BetStatus.VOID,
}


def _norm(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _money(value) -> Decimal:
    if value is None or value == "":
        return ZERO
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = (
        str(value)
        .replace("£", "")
        .replace(",", "")
        .replace("−", "-")
        .replace("–", "")
        .strip()
    )
    if not text or text in {"-", "–"}:
        return ZERO
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return ZERO


def _date(value) -> date:
    if value is None or value == "":
        return date.today()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 20000:
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except Exception:
            pass
    return parse_uk(str(value))


def _datetime(value) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)) and value > 20000:
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted
            if isinstance(converted, date):
                return datetime.combine(converted, datetime.min.time())
        except Exception:
            return None
    text = str(value).strip()
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_map(row: tuple, aliases: dict[str, set[str]]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, cell in enumerate(row):
        token = _norm(cell)
        if not token:
            continue
        for field, names in aliases.items():
            if token in names and field not in mapping:
                mapping[field] = index
    return mapping


def _find_table(ws, aliases: dict[str, set[str]], min_hits: int = 3):
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        mapping = _header_map(row, aliases)
        if len(mapping) >= min_hits:
            return row_number, mapping
    return None, {}


def _rows_after(ws, header_row: int):
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if any(cell not in (None, "") for cell in row):
            yield row


def _cell(row, mapping: dict[str, int], key: str):
    index = mapping.get(key)
    if index is None or index >= len(row):
        return None
    return row[index]


def _get_or_create_account(
    session: Session,
    name: str,
    kind: str,
    commission: Decimal | None = None,
) -> Account:
    name = _text(name)[:80]
    if not name:
        raise ValueError("Account name is missing.")
    account = session.scalars(select(Account).where(Account.name == name)).first()
    if account is None:
        account = Account(
            name=name,
            type=kind,
            commission_percent=commission if commission is not None else ZERO,
        )
        session.add(account)
        session.flush()
        return account
    if kind == AccountType.EXCHANGE:
        account.type = AccountType.EXCHANGE
    if commission is not None and account.type == AccountType.EXCHANGE:
        account.commission_percent = commission
    return account


def _offer_type(raw) -> str:
    token = _norm(raw).replace(" ", "_")
    mapping = {item.value: item.value for item in OfferType}
    mapping.update({item.name.lower(): item.value for item in OfferType})
    mapping["welcome"] = OfferType.WELCOME
    mapping["reload"] = OfferType.RELOAD
    mapping["risk free"] = OfferType.RISK_FREE
    mapping["acca insurance"] = OfferType.ACCA_INSURANCE
    mapping["extra place"] = OfferType.EXTRA_PLACE
    mapping["price boost"] = OfferType.PRICE_BOOST
    return mapping.get(_norm(raw), mapping.get(token, OfferType.OTHER))


def _bet_type(raw) -> str:
    return BET_TYPE_MAP.get(_norm(raw), BetType.OTHER)


def _status(raw) -> str:
    return STATUS_MAP.get(_norm(raw), BetStatus.PENDING)


def _default_exchange(session: Session) -> Account:
    smarkets = session.scalars(select(Account).where(Account.name == "Smarkets")).first()
    if smarkets:
        return smarkets
    first = session.scalars(
        select(Account).where(Account.type == AccountType.EXCHANGE)
    ).first()
    if first:
        return first
    return _get_or_create_account(session, "Smarkets", AccountType.EXCHANGE, Decimal("2"))


def import_workbook(
    session: Session,
    source,
    replace: bool = False,
) -> dict:
    if hasattr(source, "read"):
        workbook = load_workbook(source, data_only=True)
    else:
        workbook = load_workbook(Path(source), data_only=True)

    if replace:
        session.execute(delete(Bet))
        session.execute(delete(Transfer))
        session.execute(delete(Offer))
        session.flush()

    used: list[str] = []
    warnings: list[str] = []
    counts = {"accounts": 0, "offers": 0, "bets": 0, "transfers": 0}
    before_accounts = {
        name for name in session.scalars(select(Account.name))
    }

    sheets = {name.lower(): workbook[name] for name in workbook.sheetnames}

    if "accounts" in sheets:
        counts["accounts"] += _import_accounts(session, sheets["accounts"])
        used.append("Accounts")
    if "transfers" in sheets:
        counts["transfers"] += _import_transfers(session, sheets["transfers"])
        used.append("Transfers")
    if "offers" in sheets:
        counts["offers"] += _import_offers(session, sheets["offers"])
        used.append("Offers")
        reconcile_offer_deposits(session)

    bet_sheet = sheets.get("bets")
    if bet_sheet is None:
        for ws in workbook.worksheets:
            header_row, mapping = _find_table(ws, BET_ALIASES, min_hits=4)
            if header_row and "bookie" in mapping and "back_stake" in mapping:
                bet_sheet = ws
                used.append(ws.title)
                break
    else:
        used.append("Bets")

    if bet_sheet is not None:
        counts["bets"] += _import_bets(session, bet_sheet)
    elif "bets" not in used:
        warnings.append("No bets sheet found. Imported accounts/transfers/offers only, if present.")

    created = [
        name
        for name in session.scalars(select(Account.name))
        if name not in before_accounts
    ]
    counts["accounts"] = max(counts["accounts"], len(created))
    session.flush()
    return {"counts": counts, "sheets": used, "warnings": warnings, "new_accounts": created}


def _import_accounts(session: Session, ws) -> int:
    header_row, mapping = _find_table(ws, ACCOUNT_ALIASES, min_hits=2)
    if not mapping:
        return 0
    created = 0
    for row in _rows_after(ws, header_row):
        name = _text(_cell(row, mapping, "name"))
        if not name:
            continue
        kind_raw = _norm(_cell(row, mapping, "type"))
        kind = (
            AccountType.EXCHANGE
            if "exchange" in kind_raw
            else AccountType.BOOKIE
        )
        commission = _money(_cell(row, mapping, "commission"))
        before = session.scalars(select(Account).where(Account.name == name)).first()
        account = _get_or_create_account(session, name, kind, commission)
        if before is None:
            created += 1
        deposited = _money(_cell(row, mapping, "deposited"))
        withdrawn = _money(_cell(row, mapping, "withdrawn"))
        if deposited > 0:
            _ensure_transfer(session, account.id, TransferKind.DEPOSIT, deposited, "Imported deposit")
        if withdrawn > 0:
            _ensure_transfer(session, account.id, TransferKind.WITHDRAWAL, withdrawn, "Imported withdrawal")
    return created


def _ensure_transfer(session: Session, account_id: int, kind: str, amount: Decimal, notes: str) -> None:
    exists = session.scalars(
        select(Transfer).where(
            Transfer.account_id == account_id,
            Transfer.kind == kind,
            Transfer.amount == amount,
            Transfer.notes == notes,
        )
    ).first()
    if exists:
        return
    session.add(
        Transfer(
            account_id=account_id,
            kind=kind,
            amount=amount,
            notes=notes,
        )
    )


def _import_transfers(session: Session, ws) -> int:
    header_row, mapping = _find_table(ws, TRANSFER_ALIASES, min_hits=3)
    if not mapping:
        return 0
    count = 0
    for row in _rows_after(ws, header_row):
        name = _text(_cell(row, mapping, "account"))
        amount = _money(_cell(row, mapping, "amount"))
        if not name or amount <= 0:
            continue
        kind_raw = _norm(_cell(row, mapping, "kind"))
        kind = TransferKind.WITHDRAWAL if "withdraw" in kind_raw else (
            TransferKind.OPENING if "open" in kind_raw else TransferKind.DEPOSIT
        )
        account = session.scalars(select(Account).where(Account.name == name)).first()
        if account is None:
            account = _get_or_create_account(session, name, AccountType.BOOKIE)
        session.add(
            Transfer(
                account_id=account.id,
                kind=kind,
                amount=amount,
                date=_date(_cell(row, mapping, "date")),
                notes=_text(_cell(row, mapping, "notes")),
            )
        )
        count += 1
    return count


def _import_offers(session: Session, ws) -> int:
    header_row, mapping = _find_table(ws, OFFER_ALIASES, min_hits=2)
    if not mapping:
        return 0
    count = 0
    for row in _rows_after(ws, header_row):
        name = _text(_cell(row, mapping, "name"))
        bookie_name = _text(_cell(row, mapping, "bookie"))
        if not name or not bookie_name:
            continue
        existing = session.scalars(select(Offer).where(Offer.name == name)).first()
        if existing:
            continue
        bookie = _get_or_create_account(session, bookie_name, AccountType.BOOKIE)
        session.add(
            Offer(
                name=name,
                type=_offer_type(_cell(row, mapping, "type")),
                bookie_id=bookie.id,
                deposit_amount=_money(_cell(row, mapping, "deposited")),
                free_funds=_money(_cell(row, mapping, "free_funds")) if "free_funds" in mapping else Decimal("0"),
                notes=_text(_cell(row, mapping, "notes")),
            )
        )
        count += 1
    session.flush()
    return count


def _import_bets(session: Session, ws) -> int:
    header_row, mapping = _find_table(ws, BET_ALIASES, min_hits=4)
    if not mapping:
        return 0
    default_ex = _default_exchange(session)
    count = 0
    for row in _rows_after(ws, header_row):
        bookie_name = _text(_cell(row, mapping, "bookie"))
        if not bookie_name:
            continue
        bookie = _get_or_create_account(session, bookie_name, AccountType.BOOKIE)
        exchange_name = _text(_cell(row, mapping, "exchange"))
        exchange = (
            _get_or_create_account(session, exchange_name, AccountType.EXCHANGE)
            if exchange_name
            else default_ex
        )
        offer_name = _text(_cell(row, mapping, "offer"))
        offer_id = None
        if offer_name:
            offer = session.scalars(select(Offer).where(Offer.name == offer_name)).first()
            if offer is None:
                offer = Offer(
                    name=offer_name,
                    type=OfferType.OTHER,
                    bookie_id=bookie.id,
                )
                session.add(offer)
                session.flush()
            offer_id = offer.id
        actual = _money(_cell(row, mapping, "actual")) if "actual" in mapping else None
        status = _status(_cell(row, mapping, "status")) if "status" in mapping else (
            BetStatus.BACK_WON if actual is not None else BetStatus.PENDING
        )
        if actual is None:
            actual_val = None
        else:
            actual_val = actual
            if status == BetStatus.PENDING and actual != ZERO:
                status = BetStatus.BACK_WON
        bookie_pl = _money(_cell(row, mapping, "bookie_pl")) if "bookie_pl" in mapping else actual_val
        exchange_pl = _money(_cell(row, mapping, "exchange_pl")) if "exchange_pl" in mapping else ZERO
        back_stake = _money(_cell(row, mapping, "back_stake"))
        back_odds = _money(_cell(row, mapping, "back_odds")) or Decimal("2")
        lay_odds = _money(_cell(row, mapping, "lay_odds")) or Decimal("2")
        lay_stake = _money(_cell(row, mapping, "lay_stake"))
        commission = _money(_cell(row, mapping, "commission"))
        liability = _money(_cell(row, mapping, "liability"))
        expected = _money(_cell(row, mapping, "expected"))
        date_placed = _date(_cell(row, mapping, "date")) if "date" in mapping else date.today()
        placed_at = _datetime(_cell(row, mapping, "placed_at")) if "placed_at" in mapping else None
        if "date" not in mapping and placed_at is not None:
            date_placed = placed_at.date()
        settled_at = _datetime(_cell(row, mapping, "settled_at")) if "settled_at" in mapping else None
        session.add(
            Bet(
                offer_id=offer_id,
                date_placed=date_placed,
                placed_at=placed_at,
                event=_text(_cell(row, mapping, "event")),
                market=_text(_cell(row, mapping, "market")),
                notes=_text(_cell(row, mapping, "notes")),
                bet_type=_bet_type(_cell(row, mapping, "bet_type")),
                bookie_id=bookie.id,
                exchange_id=exchange.id,
                back_stake=back_stake,
                back_odds=back_odds,
                lay_stake=lay_stake,
                lay_odds=lay_odds,
                commission_percent=commission,
                cashback=ZERO,
                liability=liability,
                expected_profit=expected,
                expected_bookie_back=bookie_pl or ZERO,
                expected_exchange_back=exchange_pl or ZERO,
                expected_bookie_lay=bookie_pl or ZERO,
                expected_exchange_lay=exchange_pl or ZERO,
                status=status,
                actual_profit=actual_val if status != BetStatus.PENDING else None,
                actual_bookie_profit=bookie_pl if status != BetStatus.PENDING else None,
                actual_exchange_profit=exchange_pl if status != BetStatus.PENDING else None,
                settled_at=settled_at if status != BetStatus.PENDING else None,
            )
        )
        count += 1
    return count
